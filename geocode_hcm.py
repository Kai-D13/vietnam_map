"""
Geocode addresses in "long lat hcm.xlsx" using Vietmap API.
Adds/updates 'latitude' and 'longitude' columns in place (col 8, 9).
Run: python geocode_hcm.py
"""
import sys
import time
import openpyxl
import requests

# Fix Windows console encoding for Vietnamese characters
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

API_KEY   = "e3464a9335a846e985861bdf43fd8700201a93af28006a40"
BASE_URL  = "https://maps.vietmap.vn/api/search"
FILE_PATH = "long lat hcm.xlsx"

# Focus on HCM city centre to bias results
HCM_LAT = 10.762622
HCM_LON = 106.660172

# Columns (1-indexed)
COL_ADDRESS  = 4   # address
COL_LAT      = 8   # latitude  (will be created if not exists)
COL_LNG      = 9   # longitude


def geocode(address: str) -> tuple[float, float] | None:
    params = {
        "api-version": "1.1",
        "apikey":      API_KEY,
        "text":        address,
        "focus.point.lat": HCM_LAT,
        "focus.point.lon": HCM_LON,
    }
    try:
        r = requests.get(BASE_URL, params=params, timeout=10)
        r.raise_for_status()
        features = r.json().get("data", {}).get("features", [])
        if features:
            coords = features[0]["geometry"]["coordinates"]
            return float(coords[1]), float(coords[0])   # lat, lng
    except Exception as e:
        print(f"    ERROR: {e}")
    return None


def main():
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active

    # Write headers for new columns if empty
    if ws.cell(1, COL_LAT).value is None:
        ws.cell(1, COL_LAT).value = "latitude"
    if ws.cell(1, COL_LNG).value is None:
        ws.cell(1, COL_LNG).value = "longitude"

    total   = ws.max_row - 1
    success = 0
    skipped = 0
    failed  = 0

    for row in range(2, ws.max_row + 1):
        address = ws.cell(row, COL_ADDRESS).value
        existing_lat = ws.cell(row, COL_LAT).value

        if not address:
            skipped += 1
            continue

        # Skip if already geocoded
        if existing_lat is not None:
            skipped += 1
            continue

        print(f"[{row-1}/{total}] {address[:60]}", end=" ... ", flush=True)
        result = geocode(str(address))

        if result:
            lat, lng = result
            ws.cell(row, COL_LAT).value = lat
            ws.cell(row, COL_LNG).value = lng
            print(f"OK  ({lat:.6f}, {lng:.6f})")
            success += 1
        else:
            ws.cell(row, COL_LAT).value = None
            ws.cell(row, COL_LNG).value = None
            print("FAILED")
            failed += 1

        # Save every 50 rows to avoid losing progress
        if (row - 1) % 50 == 0:
            wb.save(FILE_PATH)
            print(f"  --- Saved checkpoint at row {row} ---")

        time.sleep(0.15)  # ~6-7 req/s, well within rate limit

    wb.save(FILE_PATH)
    print(f"\nDone! Success: {success} | Skipped/cached: {skipped} | Failed: {failed}")


if __name__ == "__main__":
    main()
