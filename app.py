import os
import json
import math
import time
import unicodedata
import re
import openpyxl
import requests
from dotenv import load_dotenv
from supabase import create_client as _sb_create
from flask import Flask, jsonify, render_template, request

load_dotenv()

app = Flask(__name__)

SERVICES_KEY    = os.environ.get("VIETMAP_SERVICES_KEY", "6ad21129f9a078c7ae857d4b3494c4ec426cc288368a6943")
TILEMAP_KEY     = os.environ.get("VIETMAP_TILEMAP_KEY",  "15df1a01dfcd57afa42aa1813597f6b41113fde321954745")
AREA_EXCEL_PATH = os.path.join(os.path.dirname(__file__), "area_vietnam.xlsx")
CACHE_PATH      = os.path.join(os.path.dirname(__file__), "geocode_cache.json")
DISTRICTS_PATH  = os.path.join(os.path.dirname(__file__), "data", "districts.geojson")
GRDP_PATH       = os.path.join(os.path.dirname(__file__), "vietnam_grdp.json")
MAPPING_PATH    = os.path.join(os.path.dirname(__file__), "mapping.xlsx")
SUPABASE_URL    = os.environ.get("SUPABASE_URL", "https://stqihecpcipszapcavux.supabase.co")
SUPABASE_KEY    = os.environ.get("SUPABASE_SERVICE_KEY", "")

DEPARTERS = [
    {"name": "Kho Cần Thơ",    "lat": 10.0039, "lng": 105.7725},
    {"name": "Kho Bình Dương", "lat": 11.0827, "lng": 106.6784},
    {"name": "Kho Bắc Ninh",  "lat": 21.1235, "lng": 105.9752},
]

HCM_AVG_SPEED_KMH = 25

_area_data     = None
_cache_mtime   = None
_districts_data = None
_grdp_data     = None
_carriers_data = None


def _cache_file_mtime():
    try:
        return os.path.getmtime(CACHE_PATH)
    except OSError:
        return None


def _normalize_addr(s):
    """Strip diacritics and collapse whitespace for fuzzy address matching."""
    s = (s or '').strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'\s+', ' ', s)
    return s.lower().strip()


def load_areas():
    """Load and cache area data. Reloads only when geocode_cache.json changes."""
    global _area_data, _cache_mtime

    current_mtime = _cache_file_mtime()
    if _area_data is not None and current_mtime == _cache_mtime:
        return _area_data

    t0 = time.time()

    geo = {}
    if os.path.exists(CACHE_PATH):
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            geo = json.load(f)

    wb = openpyxl.load_workbook(AREA_EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active
    areas = []
    # iter_rows is O(n) — never use ws.cell(r,c) in read_only mode (it's O(n²))
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        ward_name     = row[0]
        district_name = row[1]
        province_name = row[2]
        fc_code       = row[3]
        sales_region  = row[4]
        address       = row[5]
        total_order   = row[6]
        total_package = row[7]
        total_gmv     = row[8]
        total_cod     = row[9]
        grdp          = row[10]

        entry = geo.get(str(r_idx))
        if entry and entry.get("lat") is not None:
            lat          = entry["lat"]
            lng          = entry["lng"]
            locality_gid = entry.get("locality_gid", "")
        else:
            # Fallback: lat/lng written directly to xlsx (after --merge)
            lat_cell = row[11] if len(row) > 11 else None
            lng_cell = row[12] if len(row) > 12 else None
            if lat_cell is None or lng_cell is None:
                continue
            try:
                lat = float(lat_cell)
                lng = float(lng_cell)
            except (ValueError, TypeError):
                continue
            locality_gid = str(row[13] if len(row) > 13 else "") or ""

        areas.append({
            "ward":          str(ward_name or ""),
            "district":      str(district_name or ""),
            "province":      str(province_name or ""),
            "fc_code":       str(fc_code or ""),
            "sales_region":  str(sales_region or ""),
            "address":       str(address or ""),
            "total_order":   float(total_order) if total_order else 0,
            "total_package": float(total_package) if total_package else 0,
            "total_gmv":     float(total_gmv) if total_gmv else 0,
            "total_cod":     float(total_cod) if total_cod else 0,
            "grdp":          float(grdp) if grdp else 0,
            "locality_gid":  locality_gid,
            "lat":           lat,
            "lng":           lng,
        })

    wb.close()
    _area_data   = areas
    _cache_mtime = current_mtime
    print(f"[load_areas] Loaded {len(areas)} areas in {round(time.time()-t0, 2)}s")
    return areas


def load_carriers():
    """Load carrier data from mapping.xlsx, match coverage areas, apply Supabase overrides."""
    global _carriers_data
    if _carriers_data is not None:
        return _carriers_data

    t0 = time.time()

    # Build address lookup from area data (exact + normalized)
    areas = load_areas()
    area_addr_exact = {a['address']: a['address'] for a in areas if a.get('address')}
    area_addr_norm  = {_normalize_addr(a['address']): a['address'] for a in areas if a.get('address')}

    wb = openpyxl.load_workbook(MAPPING_PATH, read_only=True, data_only=True)

    # data_area columns: 0=province_name, 1=district_name, 2=ward_name, 3=address, 4=carrier_name
    ws_area = wb['data_area']
    carrier_area_map = {}
    unmatched = 0
    for row in ws_area.iter_rows(min_row=2, values_only=True):
        carrier = str(row[4] or '').strip()
        addr    = str(row[3] or '').strip()
        if not carrier or not addr:
            continue
        if carrier not in carrier_area_map:
            carrier_area_map[carrier] = []
        if addr in area_addr_exact:
            carrier_area_map[carrier].append(addr)
        else:
            norm = _normalize_addr(addr)
            if norm in area_addr_norm:
                carrier_area_map[carrier].append(area_addr_norm[norm])
            else:
                unmatched += 1

    if unmatched:
        print(f"[load_carriers] {unmatched} unmatched addresses (likely admin boundary renames)")

    # data_hub columns: 0=carrier_name, 1=hub_name, 2=address, 3=ward_name, 4=district_name,
    #                   5=province_name, 6=destination, 7=toa_do, 8=departer, 9=_latitude, 10=_longitude
    ws_hub = wb['data_hub']
    carriers = []
    for row in ws_hub.iter_rows(min_row=2, values_only=True):
        carrier_name = str(row[0] or '').strip()
        hub_name     = str(row[1] or '').strip()
        address      = str(row[2] or '').strip()
        departer     = str(row[8] or '').strip()
        lat          = row[9]
        lng          = row[10]
        if not carrier_name or lat is None or lng is None:
            continue
        try:
            lat = float(lat)
            lng = float(lng)
        except (ValueError, TypeError):
            continue
        coverage = carrier_area_map.get(carrier_name, [])
        carriers.append({
            "carrier_name":       carrier_name,
            "hub_name":           hub_name,
            "address":            address,
            "departer":           departer,
            "lat":                lat,
            "lng":                lng,
            "coverage_count":     len(coverage),
            "coverage_addresses": coverage,
        })
    wb.close()

    # Apply Supabase overrides (lat/lng edits and soft-deletes)
    try:
        overrides    = get_supabase().table("carrier_overrides").select("*").execute().data
        override_map = {o['carrier_name']: o for o in overrides}
        result = []
        for c in carriers:
            ov = override_map.get(c['carrier_name'])
            if ov:
                if ov.get('hidden'):
                    continue
                if ov.get('lat') is not None:
                    c['lat'] = ov['lat']
                if ov.get('lng') is not None:
                    c['lng'] = ov['lng']
            result.append(c)
        carriers = result
    except Exception as e:
        print(f"[load_carriers] Supabase override fetch failed: {e}")

    _carriers_data = carriers
    print(f"[load_carriers] Loaded {len(carriers)} carriers in {round(time.time()-t0, 2)}s")
    return carriers


@app.route("/")
def index():
    return render_template("map.html", tilemap_key=TILEMAP_KEY)


@app.route("/api/areas")
def api_areas():
    areas = load_areas()
    return jsonify({"areas": areas, "total": len(areas)})


@app.route("/api/geocode_status")
def api_geocode_status():
    if not os.path.exists(CACHE_PATH):
        return jsonify({"cached": 0, "ok": 0, "total": 10599, "done": False})
    with open(CACHE_PATH, "r", encoding="utf-8") as f:
        geo = json.load(f)
    cached = len(geo)
    ok     = sum(1 for v in geo.values() if v.get("lat") is not None)
    return jsonify({"cached": cached, "ok": ok, "total": 10599, "done": cached >= 10599})


@app.route("/api/grdp")
def api_grdp():
    global _grdp_data
    if _grdp_data is None:
        if not os.path.exists(GRDP_PATH):
            return jsonify([])
        with open(GRDP_PATH, "r", encoding="utf-8") as f:
            _grdp_data = json.load(f)
    return jsonify(_grdp_data)


@app.route("/api/route")
def api_route():
    try:
        lat1 = float(request.args["lat1"])
        lng1 = float(request.args["lng1"])
        lat2 = float(request.args["lat2"])
        lng2 = float(request.args["lng2"])
    except (KeyError, ValueError):
        return jsonify({"error": "Missing or invalid lat1/lng1/lat2/lng2"}), 400

    params = [
        ("apikey",         SERVICES_KEY),
        ("point",          f"{lat1},{lng1}"),
        ("point",          f"{lat2},{lng2}"),
        ("vehicle",        "car"),
        ("points_encoded", "true"),
        ("locale",         "vi"),
    ]
    try:
        r = requests.get("https://maps.vietmap.vn/api/route", params=params, timeout=15)
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        return jsonify({"error": str(e)}), 502

    if "paths" not in data or not data["paths"]:
        return jsonify({"error": "No route found"}), 404

    path         = data["paths"][0]
    distance_km  = round(path.get("distance", 0) / 1000, 1)
    duration_min = round(distance_km / HCM_AVG_SPEED_KMH * 60)

    return jsonify({
        "distance_km":  distance_km,
        "duration_min": duration_min,
        "points":       path.get("points", ""),
    })


@app.route("/api/boundaries/districts")
def api_boundaries_districts():
    global _districts_data
    if _districts_data is None:
        if not os.path.exists(DISTRICTS_PATH):
            return jsonify({"error": "districts.geojson not found — run download_districts.py"}), 404
        with open(DISTRICTS_PATH, "r", encoding="utf-8") as f:
            _districts_data = json.load(f)
    return jsonify(_districts_data)


# ── Carriers ──────────────────────────────────────────────────────────────────

@app.route("/api/carriers")
def api_carriers():
    carriers = load_carriers()
    slim = [{k: v for k, v in c.items() if k != 'coverage_addresses'}
            for c in carriers]
    slim.append({"_departers": DEPARTERS})
    return jsonify(slim)


@app.route("/api/carriers/<path:carrier_name>/coverage")
def api_carrier_coverage(carrier_name):
    carriers = load_carriers()
    for c in carriers:
        if c['carrier_name'] == carrier_name:
            return jsonify({"addresses": c['coverage_addresses']})
    return jsonify({"error": "not found"}), 404


@app.route("/api/carriers/<path:carrier_name>", methods=["PUT"])
def api_carriers_put(carrier_name):
    global _carriers_data
    data = request.get_json()
    if not data:
        return jsonify({"error": "no data"}), 400
    update = {"carrier_name": carrier_name}
    if "lat" in data:
        update["lat"] = float(data["lat"])
    if "lng" in data:
        update["lng"] = float(data["lng"])
    try:
        get_supabase().table("carrier_overrides").upsert(update).execute()
        _carriers_data = None
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/carriers/<path:carrier_name>", methods=["DELETE"])
def api_carriers_delete(carrier_name):
    global _carriers_data
    try:
        get_supabase().table("carrier_overrides").upsert({
            "carrier_name": carrier_name,
            "hidden": True,
        }).execute()
        _carriers_data = None
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Haversine distance ────────────────────────────────────────────────────────

def haversine_km(lat1, lng1, lat2, lng2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2))
         * math.sin(dlng / 2) ** 2)
    return R * 2 * math.asin(math.sqrt(a))


# ── Supabase client (lazy init) ───────────────────────────────────────────────

_supabase = None

def get_supabase():
    global _supabase
    if _supabase is None:
        _supabase = _sb_create(SUPABASE_URL, SUPABASE_KEY)
    return _supabase


# ── Hub CRUD via Supabase ─────────────────────────────────────────────────────

@app.route("/api/hubs", methods=["GET"])
def api_hubs_get():
    try:
        res = get_supabase().table("hubs").select("*").order("created_at").execute()
        return jsonify(res.data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/hubs", methods=["POST"])
def api_hubs_post():
    data = request.get_json()
    if not data or "lat" not in data or "lng" not in data:
        return jsonify({"error": "lat and lng required"}), 400
    hub = {
        "name":      data.get("name", "Hub"),
        "lat":       float(data["lat"]),
        "lng":       float(data["lng"]),
        "radius_km": float(data.get("radius_km", 30)),
        "stats":     data.get("stats", {}),
    }
    try:
        res = get_supabase().table("hubs").insert(hub).execute()
        return jsonify(res.data[0]), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/hubs/<hub_id>", methods=["DELETE"])
def api_hubs_delete(hub_id):
    try:
        get_supabase().table("hubs").delete().eq("id", hub_id).execute()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Geofence ──────────────────────────────────────────────────────────────────

@app.route("/api/geofence", methods=["POST"])
def api_geofence():
    data = request.get_json()
    if not data or "lat" not in data or "lng" not in data:
        return jsonify({"error": "lat and lng required"}), 400
    try:
        hub_lat   = float(data["lat"])
        hub_lng   = float(data["lng"])
        radius_km = float(data.get("radius_km", 30))
    except (ValueError, TypeError):
        return jsonify({"error": "Invalid lat/lng/radius_km"}), 400

    areas   = load_areas()
    inside  = []
    outside = []
    stats   = {"total_order": 0, "total_package": 0, "total_gmv": 0, "total_cod": 0}

    for a in areas:
        key = f"{a['ward']}|{a['district']}|{a['province']}"
        d   = haversine_km(hub_lat, hub_lng, a["lat"], a["lng"])
        if d <= radius_km:
            inside.append(key)
            stats["total_order"]   += a["total_order"]
            stats["total_package"] += a["total_package"]
            stats["total_gmv"]     += a["total_gmv"]
            stats["total_cod"]     += a["total_cod"]
        else:
            outside.append(key)

    return jsonify({"inside": inside, "outside": outside, "stats": stats})


if __name__ == "__main__":
    print("Starting Area Map server at http://localhost:5000")
    app.run(debug=True, port=5000)
