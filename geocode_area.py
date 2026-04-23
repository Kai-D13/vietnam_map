"""
Geocode addresses in "area_vietnam.xlsx" using Vietmap Search API v1.1.

Progress is saved to geocode_cache.json (avoids xlsx file-lock issues when
Excel has area_vietnam.xlsx open). After all rows are done, run with --merge
to write lat/lng/locality_gid back into area_vietnam.xlsx.

Usage:
  python geocode_area.py          # geocode, save progress to JSON
  python geocode_area.py --merge  # merge JSON cache into area_vietnam.xlsx

~10,599 rows × 0.15s ≈ 26 minutes. Resumable — skips rows already in cache.

Note: VietMap returns [0,0] for ward admin features (no centroid). We skip
those and take the first feature with valid coordinates — typically a public
building within the ward (police station, health centre, etc.).
locality_gid is saved for future use with the VietMap Boundary API.
"""
import sys
import json
import time
import argparse
import openpyxl
import requests

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

API_KEY    = "6ad21129f9a078c7ae857d4b3494c4ec426cc288368a6943"
BASE_URL   = "https://maps.vietmap.vn/api/search"
FILE_PATH  = "area_vietnam.xlsx"
CACHE_PATH = "geocode_cache.json"

VN_LAT, VN_LON = 16.0, 107.5

# Excel columns (1-indexed)
COL_ADDRESS      = 6
COL_LAT          = 10
COL_LNG          = 11
COL_LOCALITY_GID = 12


# ── Geocode ────────────────────────────────────────────────────────────────────
def geocode(address: str) -> tuple[float, float, str] | None:
    """Returns (lat, lng, locality_gid) or None.

    Skips VietMap features with [0,0] coordinates (ward admin areas have no
    centroid in this API). Returns first feature with valid coordinates.
    """
    params = {
        "api-version":     "1.1",
        "apikey":          API_KEY,
        "text":            address,
        "focus.point.lat": VN_LAT,
        "focus.point.lon": VN_LON,
    }
    try:
        r = requests.get(BASE_URL, params=params, timeout=10)
        r.raise_for_status()
        features = r.json().get("data", {}).get("features", [])
        locality_gid = ""
        for feat in features:
            coords  = feat["geometry"]["coordinates"]
            lng_val = float(coords[0])
            lat_val = float(coords[1])
            if not locality_gid:
                locality_gid = feat["properties"].get("locality_gid", "")
            if lng_val == 0.0 and lat_val == 0.0:
                continue
            if not locality_gid:
                locality_gid = feat["properties"].get("locality_gid", "")
            return lat_val, lng_val, locality_gid
    except Exception as e:
        print(f"    ERROR: {e}")
    return None


# ── Load / save JSON cache ─────────────────────────────────────────────────────
def load_cache() -> dict:
    try:
        with open(CACHE_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}


def save_cache(cache: dict):
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


# ── Main geocode loop ──────────────────────────────────────────────────────────
def run_geocode():
    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb.active

    cache   = load_cache()
    total   = ws.max_row - 1
    success = 0
    skipped = 0
    failed  = 0

    for row in range(2, ws.max_row + 1):
        address = ws.cell(row, COL_ADDRESS).value
        key     = str(row)

        if not address:
            skipped += 1
            continue

        if key in cache:
            skipped += 1
            continue

        print(f"[{row - 1}/{total}] {str(address)[:70]}", end=" ... ", flush=True)
        result = geocode(str(address))

        if result:
            lat, lng, locality_gid = result
            cache[key] = {"lat": lat, "lng": lng, "locality_gid": locality_gid}
            print(f"OK  ({lat:.6f}, {lng:.6f})")
            success += 1
        else:
            cache[key] = {"lat": None, "lng": None, "locality_gid": None}
            print("FAILED")
            failed += 1

        # Save cache every 50 rows (JSON — no file-lock issues)
        if (row - 1) % 50 == 0:
            save_cache(cache)
            pct = round((row - 1) / total * 100, 1)
            print(f"  --- Cache saved [{pct}%] {success} ok / {failed} failed ---")

        time.sleep(0.15)

    wb.close()
    save_cache(cache)
    geocoded = sum(1 for v in cache.values() if v.get("lat") is not None)
    print(f"\nDone! Success: {success} | Skipped/cached: {skipped} | Failed: {failed}")
    print(f"Total geocoded in cache: {geocoded} / {total}")
    print(f"\nRun 'python geocode_area.py --merge' to write back to {FILE_PATH}")


# ── Merge cache → xlsx ─────────────────────────────────────────────────────────
def run_merge():
    cache = load_cache()
    if not cache:
        print("No cache found. Run geocoding first.")
        return

    print(f"Loading {FILE_PATH} ...")
    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active

    # Write headers if missing
    if ws.cell(1, COL_LAT).value is None:
        ws.cell(1, COL_LAT).value = "lat"
    if ws.cell(1, COL_LNG).value is None:
        ws.cell(1, COL_LNG).value = "lng"
    if ws.cell(1, COL_LOCALITY_GID).value is None:
        ws.cell(1, COL_LOCALITY_GID).value = "locality_gid"

    merged = 0
    for row in range(2, ws.max_row + 1):
        entry = cache.get(str(row))
        if not entry:
            continue
        ws.cell(row, COL_LAT).value          = entry.get("lat")
        ws.cell(row, COL_LNG).value          = entry.get("lng")
        ws.cell(row, COL_LOCALITY_GID).value = entry.get("locality_gid")
        if entry.get("lat") is not None:
            merged += 1

    print(f"Saving {FILE_PATH} ...")
    wb.save(FILE_PATH)
    print(f"Done. {merged} rows written with coordinates.")


# ── Retry failed rows ──────────────────────────────────────────────────────────
def run_retry_failed():
    cache = load_cache()
    failed_keys = {k for k, v in cache.items() if v.get("lat") is None}
    if not failed_keys:
        print("No failed rows to retry.")
        return

    print(f"Retrying {len(failed_keys)} failed rows...")

    wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
    ws = wb.active

    # O(n) scan — collect row data for failed keys only
    row_data = {}
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if str(r_idx) in failed_keys:
            row_data[str(r_idx)] = row
    wb.close()

    retried = fixed = 0
    for key, row in sorted(row_data.items(), key=lambda x: int(x[0])):
        ward_name     = str(row[0] or "")
        district_name = str(row[1] or "")
        province_name = str(row[2] or "")

        for fallback in [
            f"{district_name}, {province_name}",
            f"{ward_name}, {province_name}",
        ]:
            print(f"[retry {key}] {fallback[:70]}", end=" ... ", flush=True)
            result = geocode(fallback)
            if result:
                lat, lng, locality_gid = result
                cache[key] = {"lat": lat, "lng": lng, "locality_gid": locality_gid}
                print(f"OK  ({lat:.6f}, {lng:.6f})")
                fixed += 1
                break
            print("FAILED")
        else:
            cache[key] = {"lat": None, "lng": None, "locality_gid": None}

        retried += 1
        time.sleep(0.15)

    save_cache(cache)
    print(f"\nRetried {retried} rows, fixed {fixed}.")


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--merge",        action="store_true", help="Merge geocode_cache.json into area_vietnam.xlsx")
    parser.add_argument("--retry-failed", action="store_true", help="Retry rows where geocoding returned null")
    args = parser.parse_args()

    if args.merge:
        run_merge()
    elif args.retry_failed:
        run_retry_failed()
    else:
        run_geocode()
